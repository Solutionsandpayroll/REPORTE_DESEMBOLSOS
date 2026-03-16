"""
generador_plano/utils.py
Genera planos bancarios usando las plantillas reales.
Cada plantilla tiene su propia lógica de datos — no se mezclan.

SANTANDER  → texto en todas las celdas de datos
BANCOLOMBIA → números enteros en: col A (tipo doc), col D (tipo transaccion=37),
               col E (código banco numérico), col L (fecha DDMMYYYY)
               Todos los números con alineación DERECHA para no quedar en rojo
"""
import os
import zipfile
import urllib.request
from io import BytesIO
from datetime import datetime
from copy import copy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ── Plantillas desde GitHub Raw (funciona en Vercel sin filesystem) ──────────
_GITHUB_RAW = "https://raw.githubusercontent.com/Solutionsandpayroll/REPORTE_DESEMBOLSOS/main/app/media"
TPL_S_URL   = f"{_GITHUB_RAW}/plantilla_plano_santander.xlsm"
TPL_B_URL   = f"{_GITHUB_RAW}/plantilla_plano_bancolombia.xlsx"

def _descargar(url: str) -> BytesIO:
    """Descarga un archivo desde GitHub Raw y lo devuelve como BytesIO."""
    with urllib.request.urlopen(url) as r:
        return BytesIO(r.read())

# ── Tipo de documento ────────────────────────────────────────────────────────
# Santander usa el texto completo; Bancolombia usa el número entero
TIPO_DOC = {
    'CC':  ('01 - CEDULA CIUDADANIA (CC)',           1),
    'CE':  ('02 - CEDULA DE EXTRANJERIA (CE)',        2),
    'NT':  ('03 - NIT (NT)',                          3),
    'NIT': ('03 - NIT (NT)',                          3),
    'TI':  ('04 - TARJETA DE IDENTIDAD (TI)',         4),
    'P':   ('05 - PASAPORTE (P)',                     5),
    'PT':  ('05 - PASAPORTE (P)',                     5),
    'PA':  ('05 - PASAPORTE (P)',                     5),
    'NE':  ('06 - NII DE ESTABLECIMIENTO (NE)',       6),
    'RC':  ('07 - REGISTRO CIVIL (RC)',               7),
    'PJ':  ('08 - PASADO JUDICIAL (PJ)',              8),
    'TP':  ('09 - TARJETA PROFESIONAL (TP)',          9),
    'SE':  ('10 - SOCIEDAD EXTRAJERA (SE)',           10),
    'BC':  ('11 - BANCO CORRESPONSAL (BC)',           11),
    'CS':  ('12 - CARNET SEGURIDAD SOCIAL (CS)',      12),
    'LC':  ('13 - LICENCIA DE CONDUCCION (LC)',       13),
    'LM':  ('14 - LIBRETA MILITAR (LM)',              14),
    'NM':  ('15 - NIT MENORES (NM)',                  15),
    'NP':  ('16 - NIT PERSONA NATURAL (NP)',          16),
    'NU':  ('17 - NRO UNICO IDENTF PERSONAL (NU)',    17),
}

# ── Bancos ────────────────────────────────────────────────────────────────────
# Cada entrada: (keyword, código_santander_texto, código_bancolombia_int, nombre_bancolombia)
BANCOS = [
    ('BANCAMIA',               '0059 - BANCAMIA S.A.',                             1059, 'BANCO DE LAS MICROFINANZAS - BANCAMIA S.A.'),
    ('BANCO AGRARIO',          '0040 - BANCO AGRARIO',                             1040, 'BANCO AGRARIO'),
    ('AV VILLAS',              '0052 - BANCO AV VILLAS',                           1052, 'BANCO AV VILLAS'),
    ('BCSC',                   '0032 - BANCO BCSC SA',                             1032, 'BANCO CAJA SOCIAL BCSC SA'),
    ('CAJA SOCIAL',            '0032 - BANCO BCSC SA',                             1032, 'BANCO CAJA SOCIAL BCSC SA'),
    ('BTG PACTUAL',            '0805 - BANCO BTG PACTUAL',                         1805, 'BANCO BTG PACTUAL'),
    ('CITIBANK',               '0009 - BANCO CITIBANK',                            1009, 'CITIBANK'),
    ('COOPCENTRAL',            '0066 - BANCO COOPCENTRAL',                         1066, 'BANCO COOPERATIVO COOPCENTRAL'),
    ('CREDIFINANCIERA',        '0558 - BANCO CREDIFINANCIERA',                     1558, 'BANCO CREDIFINANCIERA SA.'),
    ('DAVIVIENDA',             '0051 - BANCO DAVIVIENDA S.A.',                     1051, 'BANCO DAVIVIENDA SA'),
    ('BANCO DE BOGOT',         '0001 - BANCO DE BOGOTA',                           1001, 'BANCO DE BOGOTA'),
    ('BANCO DE OCCIDENTE',     '0023 - BANCO DE OCCIDENTE',                        1023, 'BANCO DE OCCIDENTE'),
    ('FALABELLA',              '0062 - BANCO FALABELLA S.A.',                      1062, 'BANCO FALABELLA S.A.'),
    ('FINANDINA',              '0063 - BANCO FINANDINA',                           1063, 'BANCO FINANDINA S.A.'),
    ('GNB SUDAMERIS',          '0012 - BANCO GNB SUDAMERIS',                       1012, 'BANCO GNB SUDAMERIS'),
    ('HSBC',                   '0010 - BANCO HSBC COLOMBIA S.A.',                  None, None),
    ('ITAU',                   '0006 - BANCO ITAU CORPBANCA COLOMBIA',             1014, 'ITAU'),
    ('JP MORGAN',              '0071 - BANCO JP MORGAN',                           1071, 'BANCO J.P. MORGAN COLOMBIA S.A.'),
    ('MUNDO MUJER',            '0047 - BANCO MUNDO MUJER',                         1047, 'BANCO MUNDO MUJER'),
    ('PICHINCHA',              '0060 - BANCO PICHINCHA S.A.',                      1060, 'BANCO PICHINCHA'),
    ('BANCO POPULAR',          '0002 - BANCO POPULAR',                             1002, 'BANCO POPULAR'),
    ('PROCREDIT',              '0058 - BANCO PROCREDIT COLOMBIA S.A.',             None, None),
    ('SANTANDER DE NEGOCIOS',  '0065 - BANCO SANTANDER DE NEGOCIOS COLOMBIA S.A', 1065, 'BANCO SANTANDER DE NEGOCIOS COLOMBIA S.A'),
    ('SERFINANZA',             '0069 - BANCO SERFINANZA S.A.',                     1069, 'BANCO SERFINANZA S.A'),
    ('BANCOLDEX',              '0031 - BANCOLDEX',                                 1031, 'BANCOLDEX S.A.'),
    ('BANCOLOMBIA',            '0007 - BANCOLOMBIA',                               1007, 'BANCOLOMBIA'),
    ('BANCOOMEVA',             '0061 - BANCOOMEVA',                                1061, 'BANCOOMEVA'),
    ('BBVA',                   '0013 - BBVA COLOMBIA',                             1013, 'BBVA COLOMBIA'),
    ('COLTEFINANCIERA',        '0370 - COLTEFINANCIERA S.A',                       1370, 'COLTEFINANCIERA S.A'),
    ('CONFIAR',                '0292 - CONFIAR COOPERTIVA FINANCIERA',             1292, 'CONFIAR COOPERATIVA FINANCIERA'),
    ('COOFINEP',               '0291 - COOFINEP COOPERATIVA FINANCIERA',           1291, 'COOFINEP COOPERATIVA FINANCIERA'),
    ('COOPERATIVA FINANCIERA ANTIOQUIA', '0283 - COOPERTIVA FINANCIERA ANTIOQUIA', 1283, 'COOPERATIVA FINANCIERA DE ANTIOQUIA'),
    ('COOTRAFA',               '0289 - COOTRAFA COOPERATIVA FINANCIERA',           1289, 'COOTRAFA COOPERATIVA FINANCIERA'),
    ('DAVIPLATA',              '0551 - DAVIPLATA',                                 1551, 'DAVIPLATA'),
    ('JURISCOOP',              '0121 - FINANCIERA JURISCOOP',                      1121, 'FINANCIERA JURISCOOP S.A. COMPAÑIA DE FINANCIAMIENTO'),
    ('IRIS',                   '0637 - IRIS',                                      1637, 'IRIS'),
    ('LULO BANK',              '0070 - LULO BANK S.A.',                            1070, 'LULO BANK S.A.'),
    ('MIBANCO',                '0067 - MIBANCO SA',                                1067, 'MIBANCO S.A.'),
    ('MOVII',                  '0801 - MOVII',                                     1801, 'MOVII'),
    ('NEQUI',                  '0507 - NEQUI',                                     1507, 'NEQUI'),
    ('NU COMP',                '0809 - NU',                                        None, None),
    ('RAPPIPAY',               '0151 - RAPPIPAY',                                  1151, 'RAPPIPAY'),
    ('SCOTIABANK',             '0019 - SCOTIABANK COLPATRIA S.A',                  1019, 'SCOTIABANK COLPATRIA S.A'),
    ('COLPATRIA',              '0019 - SCOTIABANK COLPATRIA S.A',                  1019, 'SCOTIABANK COLPATRIA S.A'),
    ('BANCO W',                '0053 - BANCO W S.A.',                              1053, 'BANCO W S.A.'),
    ('UALA',                   '0804 - UALÁ',                                      1804, 'UALA'),
    ('UALÁ',                   '0804 - UALÁ',                                      1804, 'UALA'),
    ('ASOPAGOS',               '0086 - ASOPAGOS',                                  1086, 'ASOPAGOS S.A.S'),
    ('BANCO UNION',            '0303 - BANCO UNION S.A',                           1303, 'GIROS Y FINANZAS CF'),
    ('GIROS Y FINANZAS',       '0303 - BANCO UNION S.A',                           1303, 'GIROS Y FINANZAS CF'),
]

# Listas deduplicadas para los selects del frontend
_ss, _sb = set(), set()
BANCOS_SANTANDER   = []   # ['0007 - BANCOLOMBIA', ...]
BANCOS_BANCOLOMBIA = []   # [{'codigo':1007, 'nombre':'BANCOLOMBIA'}, ...]
for _, s, c, n in BANCOS:
    if s not in _ss: _ss.add(s); BANCOS_SANTANDER.append(s)
    if c and c not in _sb: _sb.add(c); BANCOS_BANCOLOMBIA.append({'codigo': c, 'nombre': n})
BANCOS_SANTANDER.sort()
BANCOS_BANCOLOMBIA.sort(key=lambda x: x['codigo'])

# Mapa código ↔ nombre PAB
PAB_COD2NOM = {b['codigo']: b['nombre'] for b in BANCOS_BANCOLOMBIA}
PAB_NOM2COD = {b['nombre']: b['codigo'] for b in BANCOS_BANCOLOMBIA}

# ── Catálogos para el formulario del encabezado Bancolombia ─────────────────
# Cada opción: {val, label} para mostrar descripción completa en el select

TIPOS_PAGO = [
    {'val': 220, 'label': '220 — Pago a Proveedores'},
    {'val': 225, 'label': '225 — Pago de Nómina'},
    {'val': 238, 'label': '238 — Pagos a Terceros'},
    {'val': 239, 'label': '239 — Abono Obligaciones con el Banco'},
    {'val': 240, 'label': '240 — Pagos Cuenta Maestra'},
    {'val': 250, 'label': '250 — Subsidios'},
    {'val': 320, 'label': '320 — Credipago a Proveedores'},
    {'val': 325, 'label': '325 — Credipago Nómina'},
    {'val': 820, 'label': '820 — Pago Nómina Efectivo'},
    {'val': 920, 'label': '920 — Pago Proveedores Efectivo'},
]

APLICACIONES = [
    {'val': 'I', 'label': 'I — Inmediata'},
    {'val': 'M', 'label': 'M — Medio día'},
    {'val': 'N', 'label': 'N — Noche'},
]

TIPOS_CTA_DEB = [
    {'val': 'S', 'label': 'S — Ahorros'},
    {'val': 'D', 'label': 'D — Corriente'},
]

# Tipo de documento beneficiario (Bancolombia usa número entero en el archivo)
TIPOS_DOC_BCOL = [
    {'val': 1, 'label': '1 — Cédula de Ciudadanía'},
    {'val': 2, 'label': '2 — Cédula de Extranjería'},
    {'val': 3, 'label': '3 — NIT'},
    {'val': 4, 'label': '4 — Tarjeta de Identidad'},
    {'val': 5, 'label': '5 — Pasaporte'},
]

# Tipo de transacción (fijo = 37, pero se muestra el catálogo completo)
TIPOS_TRANSACCION = [
    {'val': 23, 'label': '23 — Pre-notifica cuenta Corriente'},
    {'val': 25, 'label': '25 — Pago en Efectivo'},
    {'val': 27, 'label': '27 — Abono a cuenta Corriente'},
    {'val': 33, 'label': '33 — Pre-notifica cuenta Ahorros'},
    {'val': 36, 'label': '36 — Pago Cheque Gerencia'},
    {'val': 37, 'label': '37 — Abono a cuenta de Ahorros'},
]

HDR_DEFAULT = {
    'nit_pagador':         900508955,
    'tipo_pago':           225,
    'aplicacion':          'I',
    'secuencia':           'A2',
    'nro_cuenta_debitar':  16700001881,
    'tipo_cuenta_debitar': 'D',
    'descripcion':         'Reembolso',
    'fecha_aplicacion':    '',
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def _clean_doc(v):
    s = str(v).strip()
    try:    return str(int(float(s)))
    except: return s

def _buscar_banco(nombre):
    if not nombre or str(nombre).strip() in ('', 'nan', 'NaN'):
        return '', None, ''
    n = str(nombre).strip().upper()
    for kw, sant, cod, nom_b in BANCOS:
        if kw.upper() in n:
            return sant, cod, (nom_b or '')
    return str(nombre).strip(), None, ''

def _norm_cuenta(t):
    return 'CORRIENTE' if 'CORRIENTE' in str(t).upper() else 'AHORROS'

def _norm_tdoc(t):
    return TIPO_DOC.get(str(t).strip().upper(), ('01 - CEDULA CIUDADANIA (CC)', 1))

# ── Lectura del maestro ───────────────────────────────────────────────────────
def _sheet_maestro(xl, sheet):
    raw = pd.read_excel(xl, sheet_name=sheet, header=None)
    hdr = None
    for i in range(min(5, len(raw))):
        vals = [str(v).upper().strip() for v in raw.iloc[i]]
        if any('DOCUMENTO' in v or 'DOCTO IDENT' in v for v in vals):
            hdr = i; break
    if hdr is None: hdr = 1
    df = raw.iloc[hdr:].copy()
    df.columns = [str(c).strip() for c in raw.iloc[hdr]]
    df = df.iloc[1:].reset_index(drop=True)
    doc_c = banco_c = tipo_c = cta_c = tdoc_c = nom_c = email_c = cel_c = None
    for c in df.columns:
        cu = c.upper().strip()
        if cu in ('DOCUMENTO','DOCTO IDENT','# DOCUMENTO') and not doc_c:  doc_c = c
        if 'BANCO' in cu and 'NOM' not in cu and not banco_c:              banco_c = c
        if cu == 'NOM BANCO':                                               nom_c = c
        if cu in ('TIPO CUENTA','TIPO CTA') and not tipo_c:                tipo_c = c
        if cu in ('N° CUENTA','CTA BCO') and not cta_c:                    cta_c = c
        if cu in ('TIPO DOC','CLASE DOCTO') and not tdoc_c:                tdoc_c = c
        if 'MAIL' in cu and not email_c:                                   email_c = c
        if cu == 'CELULAR' and not cel_c:                                  cel_c = c
    if nom_c: banco_c = nom_c
    df['__DOC__'] = df[doc_c].apply(_clean_doc) if doc_c else ''
    return df, banco_c, tipo_c, cta_c, tdoc_c, email_c, cel_c

def leer_maestro_externo(archivo):
    xl = pd.ExcelFile(archivo)
    df, b, t, c, d, em, cel = _sheet_maestro(xl, xl.sheet_names[0])
    out = {}
    for _, row in df.iterrows():
        doc = str(row.get('__DOC__','')).strip()
        if not doc or doc in ('','nan','0'): continue
        br = str(row.get(b,'') if b else '').strip()
        s, cod, nom = _buscar_banco(br)
        out[doc] = {
            'banco_raw': br, 'banco_sant': s, 'banco_cod': cod, 'banco_nom': nom,
            'tipo_cuenta': _norm_cuenta(str(row.get(t,'') if t else '')),
            'num_cuenta': str(row.get(c,'') if c else '').strip(),
            'tipo_doc': str(row.get(d,'CC') if d else 'CC').strip(),
            'email':    str(row.get(em,'') if em else '').strip(),
            'celular':  str(row.get(cel,'') if cel else '').strip(),
        }
    return out

# ── Lectura del formato de envío ──────────────────────────────────────────────
def leer_formato_envio(archivo, maestro_externo=None):
    xl = pd.ExcelFile(archivo)
    nombres = xl.sheet_names
    h_pago  = next((n for n in ['Hoja2','S&P'] if n in nombres), nombres[0])
    h_maest = next((n for n in ['Hoja3','Hoja1'] if n in nombres), None)

    raw = pd.read_excel(xl, sheet_name=h_pago, header=None)
    fecha = consecutivo = ''
    for ri in range(min(10, len(raw))):
        for ci in range(raw.shape[1]):
            v = str(raw.iloc[ri, ci]).upper()
            if 'FECHA' in v and ci+1 < raw.shape[1]:
                vv = raw.iloc[ri, ci+1]
                if pd.notna(vv) and str(vv) not in ('nan',''):
                    try: fecha = pd.to_datetime(vv).strftime('%Y-%m-%d')
                    except: fecha = str(vv)
            if 'CONSECUTIVO' in v and ci+1 < raw.shape[1]:
                vv = raw.iloc[ri, ci+1]
                if pd.notna(vv): consecutivo = str(vv).strip()

    hdr = None
    for ri in range(len(raw)):
        vals = [str(v).upper().strip() for v in raw.iloc[ri]]
        if any('DOCUMENTO' in v or 'NOMBRE' in v for v in vals):
            hdr = ri; break
    if hdr is None: hdr = 4

    df = pd.read_excel(xl, sheet_name=h_pago, header=hdr)
    df.columns = [str(c).strip() for c in df.columns]

    def fc(*keys):
        for k in keys:
            for c in df.columns:
                if k.upper() in c.upper(): return c
        return None

    c_num  = fc('N°','NRO','NUMERO')
    c_tdoc = fc('TIPO DOCUMENTO','TIPO DOC')
    c_doc  = fc('# DOCUMENTO','DOCUMENTO')
    c_nom  = fc('NOMBRE TRABAJADOR','NOMBRE')
    c_emp  = fc('EMPRESA')
    c_val  = fc('VALOR A PAGAR','VALOR RECIBOS','VALOR REPORTADO')

    if c_num:
        df = df[pd.to_numeric(df[c_num], errors='coerce').notna()].copy()

    df_m, bm, tm, cm, dm, em_m, cel_m = None, None, None, None, None, None, None
    if h_maest:
        df_m, bm, tm, cm, dm, em_m, cel_m = _sheet_maestro(xl, h_maest)

    registros = []
    faltantes = []

    for _, fila in df.iterrows():
        num    = int(pd.to_numeric(fila.get(c_num,0), errors='coerce') or 0)
        doc    = _clean_doc(fila.get(c_doc,'')) if c_doc else ''
        tdoc   = str(fila.get(c_tdoc,'CC')).strip() if c_tdoc else 'CC'
        nombre = str(fila.get(c_nom,'')).strip() if c_nom else ''
        empresa= str(fila.get(c_emp,'')).strip() if c_emp else ''
        valor  = float(fila.get(c_val,0)) if c_val and pd.notna(fila.get(c_val,0)) else 0.0

        br = tip = cta = tdm = email = celular = ''
        b_cod = None; b_nom = ''
        found = False

        if df_m is not None:
            emp = df_m[df_m['__DOC__'] == doc]
            if not emp.empty:
                e = emp.iloc[0]
                br     = str(e.get(bm,'') if bm else '').strip()
                tip    = str(e.get(tm,'') if tm else '').strip()
                cta    = str(e.get(cm,'') if cm else '').strip()
                tdm    = str(e.get(dm,tdoc) if dm else tdoc).strip()
                email  = str(e.get(em_m,'') if em_m else '').strip()
                celular= str(e.get(cel_m,'') if cel_m else '').strip()
                found = True

        if not found and maestro_externo and doc in maestro_externo:
            m = maestro_externo[doc]
            br=m['banco_raw']; tip=m['tipo_cuenta']; cta=m['num_cuenta']; tdm=m['tipo_doc']
            email=m.get('email',''); celular=m.get('celular','')
            found = True

        if not found:
            tdm = tdoc
            faltantes.append({'numero':num,'doc':doc,'nombre':nombre,
                               'empresa':empresa,'valor':valor,'tipo_doc':tdoc})

        sant, b_cod, b_nom = _buscar_banco(br)
        td_text, td_num = _norm_tdoc(tdm)

        registros.append({
            'numero': num,
            'tipo_doc_key':      tdm.upper(),
            'tipo_doc_texto':    td_text,
            'tipo_doc_num':      td_num,
            'documento':         doc,
            'nombre':            nombre,
            'empresa':           empresa,
            'banco_raw':         br,
            'banco_sant':        sant if sant else br,
            'banco_cod':         b_cod,
            'banco_nom':         b_nom,
            'banco_mapeado':     bool(sant) and sant != br,
            'tipo_cuenta':       _norm_cuenta(tip),
            'num_cuenta':        cta,
            'valor':             valor,
            'email':             email,
            'celular':           celular,
            'doc_autorizado':    '',   # se completa en el formulario/preview
            'encontrado':        found,
        })

    return {
        'registros': registros, 'faltantes': faltantes,
        'fecha': fecha, 'consecutivo': consecutivo,
        'total': round(sum(r['valor'] for r in registros), 2),
        'cantidad': len(registros),
    }


def aplicar_manuales(datos, manuales):
    for r in datos['registros']:
        m = manuales.get(r['documento'])
        if not m: continue
        br = m.get('banco_sant','') or m.get('banco_raw','')
        sant, cod, nom = _buscar_banco(br)
        # Si el usuario eligió banco bancolombia directamente por nombre
        if m.get('banco_nom') and m['banco_nom'] in PAB_NOM2COD:
            nom = m['banco_nom']; cod = PAB_NOM2COD[nom]
        r['banco_sant']      = sant if sant else br
        r['banco_cod']       = cod
        r['banco_nom']       = nom
        r['banco_mapeado']   = bool(sant) and sant != br
        r['tipo_cuenta']     = _norm_cuenta(m.get('tipo_cuenta', r['tipo_cuenta']))
        r['num_cuenta']      = m.get('num_cuenta', r['num_cuenta'])
        r['valor']           = float(m.get('valor', r['valor']) or r['valor'])
        r['email']           = m.get('email', r.get('email',''))
        r['celular']         = m.get('celular', r.get('celular',''))
        r['doc_autorizado']  = m.get('doc_autorizado', r.get('doc_autorizado',''))
        r['tipo_transaccion']= int(m.get('tipo_transaccion', r.get('tipo_transaccion',37)) or 37)
        td_t, td_n = _norm_tdoc(m.get('tipo_doc', r['tipo_doc_key']))
        r['tipo_doc_texto']  = td_t
        r['tipo_doc_num']    = td_n
        r['tipo_doc_key']    = m.get('tipo_doc', r['tipo_doc_key']).upper()
        r['encontrado']      = True
    datos['total']     = round(sum(r['valor'] for r in datos['registros']), 2)
    datos['faltantes'] = []
    return datos


# ── Generación SANTANDER ──────────────────────────────────────────────────────
# Logo Santander + botón rojo "Generar Archivo / Generate File" están embebidos
# en xl/drawings/drawing1.xml dentro del .xlsm.
#
# Problema: openpyxl strips the VBA shape (button) when saving — el botón
# desaparece en Vercel y otros entornos serverless.
#
# Solución: inyección ZIP directa.
#   1. Generar el xlsx con openpyxl (datos correctos)
#   2. Reemplazar xl/drawings/drawing1.xml y xl/media/image1.png con
#      los originales de la plantilla → logo + botón garantizados
#
# Columnas A-I (originales) + J-L (nuevas):
#   A: Beneficiario N       General
#   B: Tipo de Documento    @
#   C: Numero de Documento  @
#   D: Banco                @
#   E: Tipo de Cuenta       General
#   F: Cuenta destino       @
#   G: Monto                #,##0.00
#   H: Valida Documento?    General
#   I: Referencia           @
#   J: Email                @  (del maestro E-MAIL)
#   K: Documento Autorizado @  (mismo valor para todos)
#   L: Celular Beneficiario @  (del maestro CELULAR)

def _xml_escape(s):
    """Escapa caracteres especiales XML."""
    return (str(s)
            .replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;')
            .replace('"', '&quot;')
            .replace("'", '&apos;'))


def generar_santander(datos, referencia=''):
    """
    Estrategia: edición XML directa dentro del ZIP original.
    - Descarga la plantilla .xlsm desde GitHub Raw (BytesIO)
    - Abre el ZIP y copia TODOS los archivos sin tocarlos
    - Solo reemplaza xl/worksheets/sheet1.xml con los nuevos datos
    - El drawing, logo, VBA, relaciones, content-types quedan 100% intactos
    """
    import re as _re

    regs = datos['registros']
    cons = datos.get('consecutivo', '')
    ref  = referencia or (f"Reembolso {cons}" if cons else "Reembolso")

    doc_auto_global = next(
        (r.get('doc_autorizado', '').strip() for r in regs if r.get('doc_autorizado', '')),
        ''
    )

    # ── Descargar plantilla original ─────────────────────────────────────────
    tpl_bytes = _descargar(TPL_S_URL)

    with zipfile.ZipFile(tpl_bytes, 'r') as zin:
        sheet1_original = zin.read('xl/worksheets/sheet1.xml').decode('utf-8')

        # ── Construir nuevas filas de datos (fila 6 en adelante) ─────────────
        # Estilos de la plantilla original (obtenidos inspeccionando el XML):
        #   s="5"  → General  (col A, E, H)
        #   s="6"  → @        (col B, C, F, I)
        #   s="11" → #,##0.00 (col G)
        # Mapeamos cada columna a su estilo original
        # Cols D (banco) y las nuevas J,K,L reusan s="6" (texto @)
        COL_STYLE = {
            'A': '5',   # Beneficiario   General
            'B': '6',   # Tipo Doc       @
            'C': '6',   # Num Doc        @
            'D': '6',   # Banco          @
            'E': '5',   # Tipo Cuenta    General
            'F': '6',   # Num Cuenta     @
            'G': '11',  # Monto          #,##0.00
            'H': '5',   # Valida Doc     General
            'I': '6',   # Referencia     @
        }
        COLS = list('ABCDEFGHI')

        def fila_xml(rn, vals_dict):
            """Genera el XML de una fila completa."""
            cells = []
            for col in COLS:
                val = vals_dict.get(col, '')
                st  = COL_STYLE[col]
                ref_cell = f'{col}{rn}'
                if val == '' or val is None:
                    cells.append(f'<c r="{ref_cell}" s="{st}"/>')
                elif col == 'G':
                    # Valor numérico
                    n = float(val) if val != '' else 0.0
                    cells.append(f'<c r="{ref_cell}" s="{st}" t="n"><v>{n}</v></c>')
                else:
                    # Texto inline (t="inlineStr") para evitar depender del sharedStrings
                    cells.append(
                        f'<c r="{ref_cell}" s="{st}" t="inlineStr">'
                        f'<is><t>{_xml_escape(val)}</t></is></c>'
                    )
            return f'<row r="{rn}" spans="1:9">{"".join(cells)}</row>'

        nuevas_filas = []
        for i, r in enumerate(regs):
            rn = 6 + i
            nuevas_filas.append(fila_xml(rn, {
                'A': f"Beneficiario {r['numero']}",
                'B': r['tipo_doc_texto'],
                'C': r['documento'],
                'D': r['banco_sant'],
                'E': r['tipo_cuenta'],
                'F': r['num_cuenta'],
                'G': r['valor'],
                'H': 'SI',
                'I': ref,
            }))

        # Filas vacías para limpiar el resto (6+len hasta 10005)
        for rn in range(6 + len(regs), 10006):
            cells = ''.join(
                f'<c r="{col}{rn}" s="{COL_STYLE[col]}"/>' for col in COLS
            )
            nuevas_filas.append(f'<row r="{rn}" spans="1:9">{cells}</row>')

        bloque_datos = ''.join(nuevas_filas)

        # ── Reconstruir sheet1.xml ────────────────────────────────────────────
        # Extraer el bloque sheetData y reemplazar solo las filas 6 en adelante
        # Filas 1-5 (encabezados y fórmulas) se conservan intactas
        sd_start = sheet1_original.find('<sheetData>')
        sd_end   = sheet1_original.find('</sheetData>') + len('</sheetData>')

        # Obtener filas 1-5 del sheetData original
        sheetdata_orig = sheet1_original[sd_start:sd_end]
        filas_1_5 = _re.findall(
            r'<row r="[1-5]".*?</row>', sheetdata_orig, _re.DOTALL
        )
        encabezados = ''.join(filas_1_5)

        # Actualizar el valor cacheado <v> de las formulas D2 y D3
        # D2 = COUNTIF -> cantidad de registros
        # D3 = SUMIFS  -> monto total
        # Excel usa este valor para mostrar el dato antes de recalcular
        cantidad = datos.get('cantidad', len(regs))
        total    = datos.get('total', sum(r['valor'] for r in regs))
        encabezados = _re.sub(
            r'(<c r="D2"[^>]*><f>[^<]*</f><v>)[^<]*(</v>)',
            r'\g<1>' + str(cantidad) + r'\g<2>',
            encabezados
        )
        encabezados = _re.sub(
            r'(<c r="D3"[^>]*><f>[^<]*</f><v>)[^<]*(</v>)',
            r'\g<1>' + str(total) + r'\g<2>',
            encabezados
        )

        nuevo_sheetdata = f'<sheetData>{encabezados}{bloque_datos}</sheetData>'
        sheet1_nuevo = (
            sheet1_original[:sd_start]
            + nuevo_sheetdata
            + sheet1_original[sd_end:]
        )

        # ── Reensamblar el ZIP copiando todo excepto sheet1.xml ───────────────
        buf_out = BytesIO()
        with zipfile.ZipFile(buf_out, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == 'xl/worksheets/sheet1.xml':
                    zout.writestr(item, sheet1_nuevo.encode('utf-8'))
                elif item.filename == 'xl/calcChain.xml':
                    # calcChain puede causar errores de recálculo; lo eliminamos
                    # Excel lo regenera automáticamente al abrir
                    pass
                else:
                    zout.writestr(item, zin.read(item.filename))

        buf_out.seek(0)
        return buf_out


# ── Generación BANCOLOMBIA ────────────────────────────────────────────────────
# Estructura:
#   Fila 1: encabezados del pagador (bold, center)
#   Fila 2: valores del pagador
#   Fila 3: encabezados de beneficiarios
#   Fila 4+: datos de beneficiarios
#
# Columnas beneficiarios (A-L):
#   A: Tipo Documento       → NÚMERO entero (1, 2, 3...)   DERECHA
#   B: Nit Beneficiario     → texto (número como string)   IZQUIERDA
#   C: Nombre Beneficiario  → texto                        IZQUIERDA
#   D: Tipo Transaccion     → NÚMERO entero (37)           DERECHA
#   E: Código Banco         → NÚMERO entero (1001, 1007)   DERECHA
#   F: No Cuenta Beneficiario → texto                      IZQUIERDA  fmt=@
#   G: Email                → texto vacío
#   H: Documento Autorizado → texto vacío
#   I: Referencia           → texto
#   J: Celular Beneficiario → texto vacío
#   K: ValorTransaccion     → NÚMERO decimal               DERECHA    fmt=#,##0.00
#   L: Fecha de aplicación  → NÚMERO entero DDMMYYYY       DERECHA
#
# IMPORTANTE: Números deben tener alineación DERECHA o quedan en ROJO en Excel
ALIN_DER = Alignment(horizontal='right',  vertical='center')
ALIN_IZQ = Alignment(horizontal='left',   vertical='center')

def generar_bancolombia(datos, hdr):
    """
    hdr = dict con claves:
      nit_pagador, tipo_pago, aplicacion, secuencia,
      nro_cuenta_debitar, tipo_cuenta_debitar, descripcion, fecha_aplicacion
    """
    regs = datos['registros']
    desc = hdr.get('descripcion','Reembolso')

    fecha_str = hdr.get('fecha_aplicacion','')
    if fecha_str:
        try:
            dt = datetime.strptime(fecha_str, '%Y-%m-%d')
            fecha_int = int(dt.strftime('%d%m%Y'))
        except:
            fecha_int = int(datetime.today().strftime('%d%m%Y'))
    else:
        fecha_int = int(datetime.today().strftime('%d%m%Y'))

    wb = load_workbook(_descargar(TPL_B_URL))
    ws = wb['FORMATOPAB']

    # Limpiar datos desde fila 4
    for rn in range(4, ws.max_row + 1):
        for ci in range(1, 13):
            ws.cell(rn, ci).value = None

    # Actualizar fila 2 (valores del pagador) con el formulario
    ws['A2'].value = int(hdr.get('nit_pagador', 900508955) or 900508955)
    ws['B2'].value = int(hdr.get('tipo_pago', 225) or 225)
    ws['C2'].value = str(hdr.get('aplicacion', 'I') or 'I').strip()
    ws['D2'].value = str(hdr.get('secuencia', 'A2') or 'A2').strip()
    ws['E2'].value = int(hdr.get('nro_cuenta_debitar', 16700001881) or 16700001881)
    ws['F2'].value = str(hdr.get('tipo_cuenta_debitar', 'D') or 'D').strip()
    ws['G2'].value = desc

    # Capturar estilos base de fila 4 (ya limpia)
    est = {ci: {'font': copy(ws.cell(4,ci).font), 'fill': copy(ws.cell(4,ci).fill),
                'border': copy(ws.cell(4,ci).border)} for ci in range(1,13)}

    # Formatos numéricos exactos de la plantilla original
    FMT = {
        1: 'General',    # Tipo Doc → número entero sin formato
        2: '@',          # Nit → texto
        3: '@',          # Nombre → texto (no 0.00E+00, eso es el original pero da error visual)
        4: 'General',    # Tipo Transaccion → número
        5: 'General',    # Código Banco → número
        6: '@',          # Cuenta → texto
        7: 'General',    # Email
        8: 'General',    # Doc Autorizado
        9: 'General',    # Referencia
        10: 'General',   # Celular
        11: '#,##0.00',  # Valor → número con decimales
        12: 'General',   # Fecha → número entero
    }
    # Alineación por columna: números a la DERECHA para evitar el rojo de Excel
    ALIN = {
        1: ALIN_DER,   # Tipo Doc (número)
        2: ALIN_IZQ,   # Nit (texto)
        3: ALIN_IZQ,   # Nombre (texto)
        4: ALIN_DER,   # Tipo Transaccion (número)
        5: ALIN_DER,   # Código Banco (número)
        6: ALIN_IZQ,   # Cuenta (texto)
        7: ALIN_IZQ,   # Email
        8: ALIN_IZQ,   # Doc Autorizado
        9: ALIN_IZQ,   # Referencia
        10: ALIN_IZQ,  # Celular
        11: ALIN_DER,  # Valor (número)
        12: ALIN_DER,  # Fecha (número)
    }

    for i, r in enumerate(regs):
        rn = 4 + i

        # Código banco: debe ser el número entero (1007), no el nombre
        banco_cod = r.get('banco_cod')
        if banco_cod is None:
            banco_cod = ''   # sin mapear → vacío

        vals = {
            1:  r['tipo_doc_num'],           # NÚMERO entero: 1, 2, 3...
            2:  r['documento'],              # texto
            3:  r['nombre'],                 # texto
            4:  r.get('tipo_transaccion',37),# NÚMERO (default 37)
            5:  banco_cod,                   # NÚMERO entero: 1007, 1051...
            6:  r['num_cuenta'],             # texto
            7:  r.get('email',''),           # Email del maestro
            8:  r.get('doc_autorizado',''),  # Documento Autorizado (editable)
            9:  desc,                        # Referencia
            10: r.get('celular',''),         # Celular del maestro
            11: r['valor'],                  # NÚMERO decimal
            12: fecha_int,                   # NÚMERO entero DDMMYYYY
        }

        for ci, val in vals.items():
            cell = ws.cell(rn, ci, val)
            cell.font      = copy(est[ci]['font'])
            cell.fill      = copy(est[ci]['fill'])
            cell.border    = copy(est[ci]['border'])
            cell.number_format = FMT[ci]
            cell.alignment = ALIN[ci]   # ← CRÍTICO: números a la derecha

    out = BytesIO(); wb.save(out); out.seek(0)
    return out